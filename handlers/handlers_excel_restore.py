"""
Импорт БД из Excel, полученного через /export_full (или /export по листу users).
Только админы. Команда /import_excel — затем отправьте .xlsx следующим сообщением.
"""
from __future__ import annotations

import asyncio
import os
import re
import tempfile
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple, Type

import openpyxl
from aiogram import F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command
from aiogram.types import Message

from bot import sql, x3
from config import ADMIN_IDS
from config_bd.utils import _naive_utc
from config_bd.models import (
    Gifts,
    Online,
    Payments,
    PaymentsCards,
    PaymentsCryptobot,
    PaymentsFkSBP,
    PaymentsPlategaCrypto,
    PaymentsStars,
    PaymentsWataCard,
    PaymentsWataSBP,
    Users,
    WhiteCounter,
)
from logging_config import logger

router = Router()

_WAITING_IMPORT_EXCEL: set[int] = set()
_WAITING_EXPORT_USERS: set[int] = set()
_WAITING_IMPORT_PAYS: set[int] = set()

# Лимит Telegram Bot API на скачивание файла ботом (getFile), не путать с лимитом отправки в чат.
TELEGRAM_BOT_MAX_DOWNLOAD_BYTES = 20 * 1024 * 1024


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", "_", s)
    return s


def _parse_datetime(val: Any) -> Optional[datetime]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None) if val.tzinfo else val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, datetime.min.time())
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _parse_date(val: Any) -> Optional[date]:
    dt = _parse_datetime(val)
    if dt is None:
        return None
    return dt.date()


def _parse_bool(val: Any) -> bool:
    if val is None or val == "":
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "да", "t", "y", "on")


def _parse_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(round(val))
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None


def _parse_bigint(val: Any) -> Optional[int]:
    v = _parse_int(val)
    return v


def _parse_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def _parse_str(val: Any, max_len: Optional[int] = None) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    if max_len is not None:
        s = s[:max_len]
    return s


def _coerce_for_column(model: Type, col_name: str, raw: Any) -> Any:
    col = model.__table__.columns.get(col_name)
    if col is None:
        return raw
    pt = getattr(col.type, "python_type", None)
    if pt is bool:
        return _parse_bool(raw)
    if pt is int:
        return _parse_int(raw)
    if pt is float:
        return _parse_float(raw)
    if pt is datetime:
        return _parse_datetime(raw)
    if pt is date:
        return _parse_date(raw)
    if pt is str:
        impl = getattr(col.type, "length", None)
        max_len = int(impl) if impl else None
        return _parse_str(raw, max_len)
    return raw


_USER_DEFAULTS: Dict[str, Any] = {
    "stamp": "",
    "is_delete": False,
    "in_panel": False,
    "is_connect": False,
    "in_chanel": False,
    "reserve_field": False,
    "field_bool_1": False,
    "field_bool_2": False,
    "field_bool_3": False,
}


def _build_user(row_map: Dict[str, Any]) -> Optional[Users]:
    """row_map: ключи как в Excel (имена колонок модели Users)."""
    lower = {_norm_header(k): v for k, v in row_map.items() if k is not None}
    by_attr: Dict[str, Any] = {}
    for c in Users.__table__.columns:
        key = c.key
        nk = _norm_header(key)
        raw = None
        if key in row_map:
            raw = row_map[key]
        elif nk in lower:
            raw = lower[nk]
        if raw is None or raw == "":
            continue
        by_attr[key] = _coerce_for_column(Users, key, raw)
    if by_attr.get("user_id") is None:
        return None
    col_keys = {c.key for c in Users.__table__.columns}
    kwargs: Dict[str, Any] = {}
    for c in Users.__table__.columns:
        key = c.key
        if key in by_attr:
            kwargs[key] = by_attr[key]
        elif key in _USER_DEFAULTS:
            kwargs[key] = _USER_DEFAULTS[key]
    if not kwargs.get("stamp"):
        kwargs["stamp"] = ""
    kwargs.pop("id", None)
    try:
        return Users(**{k: v for k, v in kwargs.items() if k in col_keys})
    except Exception:
        logger.exception("users row skip")
        return None


def _row_to_dict(headers: List[Any], row: Tuple[Any, ...]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        name = str(h).strip()
        if not name:
            continue
        val = row[i] if i < len(row) else None
        out[name] = val
    return out


def _map_payment_row(headers: List[Any], row: Tuple[Any, ...], model: Type) -> Dict[str, Any]:
    raw = _row_to_dict(headers, row)
    norm_map = {_norm_header(k): v for k, v in raw.items()}
    aliases = {
        "id": "id",
        "user_id": "user_id",
        "amount": "amount",
        "time_created": "time_created",
        "is_gift": "is_gift",
        "status": "status",
        "transaction_id": "transaction_id",
        "payload": "payload",
        "currency": "currency",
        "invoice_id": "invoice_id",
        "fk_order_id": "fk_order_id",
        "nonce": "nonce",
        "signature": "signature",
        "method": "method",
    }
    out: Dict[str, Any] = {}
    for nk, attr in aliases.items():
        if nk not in norm_map:
            continue
        if attr not in model.__table__.columns:
            continue
        out[attr] = _coerce_for_column(model, attr, norm_map[nk])
    return out


_ONLINE_ATTR_BY_HEADER = {
    "id": "online_id",
    "дата_сбора": "online_date",
    "всего_в_панели": "users_panel",
    "активны_сегодня": "users_active",
    "платных": "users_pay",
    "триальных": "users_trial",
}


def _map_online_row(headers: List[Any], row: Tuple[Any, ...]) -> Dict[str, Any]:
    raw = _row_to_dict(headers, row)
    out: Dict[str, Any] = {}
    for k, v in raw.items():
        nk = _norm_header(k)
        attr = _ONLINE_ATTR_BY_HEADER.get(nk)
        if not attr:
            continue
        out[attr] = _coerce_for_column(Online, attr, v)
    for f in ("users_panel", "users_active", "users_pay", "users_trial"):
        if f not in out or out[f] is None:
            out[f] = 0
    return out


def _iter_sheet_rows(ws: Any) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [h if h is None else str(h).strip() for h in rows[0]]
    data = [tuple(r) for r in rows[1:] if any(c is not None and c != "" for c in r)]
    return headers, data


def _parse_workbook(path: str) -> Dict[str, List[Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: Dict[str, List[Any]] = {
        "users": [],
        "payments": [],
        "payments_cards": [],
        "payments_platega_crypto": [],
        "payments_stars": [],
        "payments_cryptobot": [],
        "payments_wata_sbp": [],
        "payments_wata_card": [],
        "payments_fk_sbp": [],
        "gifts": [],
        "online": [],
        "white_counter": [],
    }

    sheets = {sn.lower(): sn for sn in wb.sheetnames}

    def get(name: str) -> Optional[str]:
        return sheets.get(name.lower())

    if get("users"):
        ws = wb[get("users")]
        headers, data = _iter_sheet_rows(ws)
        for row in data:
            d = _row_to_dict(headers, row)
            u = _build_user(d)
            if u:
                out["users"].append(u)

    def load_payments(sheet_key: str, model: Type, out_key: str):
        sn = get(sheet_key)
        if not sn:
            return
        ws = wb[sn]
        headers, data = _iter_sheet_rows(ws)
        for row in data:
            kw = _map_payment_row(headers, row, model)
            if kw.get("user_id") is None:
                continue
            if "amount" in model.__table__.columns and kw.get("amount") is None:
                kw["amount"] = 0.0 if model is PaymentsCryptobot else 0
            kw = {k: v for k, v in kw.items() if v is not None}
            try:
                out[out_key].append(model(**kw))
            except Exception:
                logger.exception("%s row skip", out_key)

    load_payments("payments_sbp", Payments, "payments")
    load_payments("payments_cards", PaymentsCards, "payments_cards")
    load_payments("payments_platega_crypto", PaymentsPlategaCrypto, "payments_platega_crypto")
    load_payments("payments_stars", PaymentsStars, "payments_stars")
    load_payments("payments_cryptobot", PaymentsCryptobot, "payments_cryptobot")
    load_payments("payments_wata_sbp", PaymentsWataSBP, "payments_wata_sbp")
    load_payments("payments_wata_card", PaymentsWataCard, "payments_wata_card")
    load_payments("payments_fk_sbp", PaymentsFkSBP, "payments_fk_sbp")

    if get("gifts"):
        ws = wb[get("gifts")]
        headers, data = _iter_sheet_rows(ws)
        for row in data:
            raw = _row_to_dict(headers, row)
            norm = {_norm_header(k): v for k, v in raw.items()}
            try:
                gid = _parse_str(norm.get("gift_id") or raw.get("gift_id"), 36)
                if not gid:
                    continue
                gv = _parse_bigint(norm.get("giver_id"))
                if gv is None:
                    continue
                out["gifts"].append(
                    Gifts(
                        gift_id=gid,
                        giver_id=gv,
                        duration=_parse_int(norm.get("duration")) or 0,
                        recepient_id=_parse_bigint(norm.get("recepient_id")),
                        white_flag=_parse_bool(norm.get("white_flag")),
                        flag=_parse_bool(norm.get("flag")),
                    )
                )
            except Exception:
                logger.exception("gifts row skip")

    if get("online"):
        ws = wb[get("online")]
        headers, data = _iter_sheet_rows(ws)
        for row in data:
            kw = _map_online_row(headers, row)
            if kw.get("online_date") is None:
                continue
            try:
                out["online"].append(Online(**kw))
            except Exception:
                logger.exception("online row skip")

    if get("white_counter"):
        ws = wb[get("white_counter")]
        headers, data = _iter_sheet_rows(ws)
        for row in data:
            kw = _map_payment_row(headers, row, WhiteCounter)
            if kw.get("user_id") is None:
                continue
            kw = {k: v for k, v in kw.items() if v is not None}
            try:
                out["white_counter"].append(WhiteCounter(**kw))
            except Exception:
                logger.exception("white_counter row skip")

    wb.close()
    return out


def _cell_by_aliases(row: Dict[str, Any], *aliases: str) -> Any:
    norm_map = {_norm_header(k): v for k, v in row.items() if k is not None}
    for a in aliases:
        na = _norm_header(a)
        if na in norm_map:
            return norm_map[na]
        if a in row:
            return row[a]
    return None


def _parse_export_users_rows(path: str) -> Tuple[List[Tuple[int, datetime, Optional[datetime], Optional[datetime]]], int]:
    """
    Лист users или первый лист: колонки user_id, created_at (или create_user), trial_at, connected_at.
    Возвращает список строк и число пропущенных (нет user_id или даты создания).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_map = {s.lower(): s for s in wb.sheetnames}
        sheet_name = sheet_map.get("users") or wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], 0
        headers = [h if h is None else str(h).strip() for h in rows[0]]
        out: List[Tuple[int, datetime, Optional[datetime], Optional[datetime]]] = []
        skipped = 0
        for r in rows[1:]:
            if not any(c is not None and str(c).strip() != "" for c in r):
                continue
            raw = _row_to_dict(headers, tuple(r))
            uid = _parse_bigint(_cell_by_aliases(raw, "user_id"))
            created = _parse_datetime(_cell_by_aliases(raw, "created_at", "create_user"))
            trial = _parse_datetime(_cell_by_aliases(raw, "trial_at"))
            conn = _parse_datetime(_cell_by_aliases(raw, "connected_at"))
            if uid is None or created is None:
                skipped += 1
                continue
            out.append((uid, created, trial, conn))
        return out, skipped
    finally:
        wb.close()


def _casual_days_by_amount(amount: int) -> Optional[int]:
    if amount == 99:
        return 7
    if amount in (149, 249):
        return 30
    if amount == 539:
        return 90
    if amount == 999:
        return 180
    return None


def _parse_import_pays_rows(path: str) -> Tuple[List[Tuple[int, int]], int]:
    """
    Лист payments_sbp (как в /export) или первый лист.
    Колонки: User ID / user_id, Amount / amount.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_map = {s.lower(): s for s in wb.sheetnames}
        sheet_name = sheet_map.get("payments_sbp") or wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], 0
        headers = [h if h is None else str(h).strip() for h in rows[0]]
        out: List[Tuple[int, int]] = []
        skipped = 0
        for r in rows[1:]:
            if not any(c is not None and str(c).strip() != "" for c in r):
                continue
            raw = _row_to_dict(headers, tuple(r))
            uid = _parse_bigint(
                _cell_by_aliases(raw, "user_id", "User ID", "user id")
            )
            amt = _parse_int(_cell_by_aliases(raw, "amount", "Amount"))
            if uid is None or amt is None:
                skipped += 1
                continue
            out.append((uid, amt))
        return out, skipped
    finally:
        wb.close()


async def _flush_import_pays_log(message: Message, lines: List[str]) -> None:
    if not lines:
        return
    chunk: List[str] = []
    size = 0
    max_chunk = 3800
    for line in lines:
        line_len = len(line) + 1
        if chunk and size + line_len > max_chunk:
            await message.answer("\n".join(chunk))
            chunk = []
            size = 0
        chunk.append(line)
        size += line_len
    if chunk:
        await message.answer("\n".join(chunk))


@router.message(Command(commands=["import_excel"]))
async def import_excel_start(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Команда только для администраторов.")
        return
    _WAITING_IMPORT_EXCEL.add(message.from_user.id)
    await message.answer(
        "📥 Отправьте файл <b>.xlsx</b> (желательно из <code>/export_full</code>) "
        "<b>следующим сообщением</b>.\n\n"
        "⚠️ Все текущие данные в таблицах users, платежах, gifts, online, white_counter "
        "будут <b>удалены</b> и заменены содержимым файла.\n\n"
        "📎 Через бота можно скачать файл до <b>20 МБ</b> (ограничение Telegram API). "
        "Больше — залейте .xlsx на сервер и выполните:\n"
        "<code>python scripts/import_export_xlsx.py /путь/к/файлу.xlsx</code>\n\n"
        "Отмена: <code>/import_excel_cancel</code>",
        parse_mode="HTML",
    )


@router.message(Command(commands=["import_excel_cancel"]))
async def import_excel_cancel(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        return
    _WAITING_IMPORT_EXCEL.discard(message.from_user.id)
    await message.answer("Импорт отменён (ожидание файла сброшено).")


@router.message(Command(commands=["export_users"]))
async def export_users_start(message: Message) -> None:
    """Импорт пользователей из .xlsx в БД (команда названа по ТЗ)."""
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Команда только для администраторов.")
        return
    _WAITING_EXPORT_USERS.add(message.from_user.id)
    await message.answer(
        "📥 Отправьте файл <b>.xlsx</b> со столбцами:\n"
        "<code>user_id</code>, <code>created_at</code> (или <code>create_user</code>), "
        "<code>trial_at</code>, <code>connected_at</code>\n\n"
        "Строки без <code>user_id</code> или даты создания будут пропущены.\n\n"
        "📎 До <b>20 МБ</b> (лимит Telegram API).\n\n"
        "Отмена: <code>/export_users_cancel</code>",
        parse_mode="HTML",
    )


@router.message(Command(commands=["export_users_cancel"]))
async def export_users_cancel(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        return
    _WAITING_EXPORT_USERS.discard(message.from_user.id)
    await message.answer("Загрузка пользователей отменена.")


@router.message(Command(commands=["import_pays"]))
async def import_pays_start(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Команда только для администраторов.")
        return
    _WAITING_IMPORT_PAYS.add(message.from_user.id)
    await message.answer(
        "💳 Отправьте <b>.xlsx</b> с платежами (лист <code>payments_sbp</code> как в "
        "<code>/export</code>): колонки <code>User ID</code>, <code>Amount</code>.\n\n"
        "Обычные суммы: 99→+7д, 149/249→+30д, 539→+90д, 999→+180д; суммы <code>1</code> и "
        "<code>399</code> обрабатываются отдельно (399 — white +30д).\n\n"
        "📎 До <b>20 МБ</b>.\n\n"
        "Отмена: <code>/import_pays_cancel</code>",
        parse_mode="HTML",
    )


@router.message(Command(commands=["import_pays_cancel"]))
async def import_pays_cancel(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        return
    _WAITING_IMPORT_PAYS.discard(message.from_user.id)
    await message.answer("Обновление подписок по файлу отменено.")


@router.message(
    F.document,
    lambda m: bool(m.from_user and m.from_user.id in _WAITING_EXPORT_USERS),
)
async def export_users_document(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        _WAITING_EXPORT_USERS.discard(message.from_user.id)
        return

    doc = message.document
    if not doc or not doc.file_name:
        await message.answer("Пришлите файл с расширением .xlsx")
        return
    if not doc.file_name.lower().endswith(".xlsx"):
        await message.answer("Нужен файл .xlsx (Excel).")
        return

    if doc.file_size is not None and doc.file_size > TELEGRAM_BOT_MAX_DOWNLOAD_BYTES:
        _WAITING_EXPORT_USERS.discard(message.from_user.id)
        await message.answer(
            "❌ Файл больше <b>20 МБ</b>.\n"
            "Скопируйте .xlsx на сервер и выполните импорт локально или разбейте файл.",
            parse_mode="HTML",
        )
        return

    _WAITING_EXPORT_USERS.discard(message.from_user.id)
    await message.answer("⏳ Загружаю и записываю пользователей…")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    try:
        await message.bot.download(doc, destination=tmp_path)
        rows, skipped_parse = await asyncio.to_thread(_parse_export_users_rows, tmp_path)
        ok = 0
        failed = 0
        for uid, created, trial_at, connected_at in rows:
            try:
                has_trial = trial_at is not None
                has_connected = connected_at is not None
                if not has_trial and not has_connected:
                    await sql.upsert_user_from_export_users_xlsx(
                        uid,
                        created,
                        in_panel=False,
                        is_connect=False,
                        subscription_end_date=None,
                        subscribtion=None,
                    )
                elif has_trial and not has_connected:
                    sub = x3.generate_client_id(uid)
                    end_dt = created + timedelta(days=7)
                    await sql.upsert_user_from_export_users_xlsx(
                        uid,
                        created,
                        in_panel=True,
                        is_connect=False,
                        subscription_end_date=end_dt,
                        subscribtion=sub,
                    )
                elif has_trial and has_connected:
                    sub = x3.generate_client_id(uid)
                    end_dt = created + timedelta(days=7)
                    await sql.upsert_user_from_export_users_xlsx(
                        uid,
                        created,
                        in_panel=True,
                        is_connect=True,
                        subscription_end_date=end_dt,
                        subscribtion=sub,
                    )
                else:
                    # connected_at задан, trial_at пуст — считаем как подключённого с триалом
                    sub = x3.generate_client_id(uid)
                    end_dt = created + timedelta(days=7)
                    await sql.upsert_user_from_export_users_xlsx(
                        uid,
                        created,
                        in_panel=True,
                        is_connect=True,
                        subscription_end_date=end_dt,
                        subscribtion=sub,
                    )
                ok += 1
            except Exception:
                failed += 1
                logger.exception("export_users row user_id=%s", uid)

        await message.answer(
            "✅ Готово.\n"
            f"Записано строк: <b>{ok}</b>\n"
            f"Ошибок при записи: <b>{failed}</b>\n"
            f"Пропущено при разборе (нет user_id или даты): <b>{skipped_parse}</b>",
            parse_mode="HTML",
        )
        logger.info(
            "Админ %s import export_users: ok=%s failed=%s skipped=%s",
            message.from_user.id,
            ok,
            failed,
            skipped_parse,
        )
    except TelegramBadRequest as e:
        err = str(e).lower()
        logger.exception("export_users Telegram error")
        if "too big" in err or "file is too large" in err:
            await message.answer(
                "❌ Telegram: файл слишком большой для скачивания ботом (~20 МБ).",
                parse_mode="HTML",
            )
        else:
            await message.answer(f"❌ Telegram: {e}")
    except Exception as e:
        logger.exception("export_users failed")
        await message.answer(f"❌ Ошибка: {e}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@router.message(
    F.document,
    lambda m: bool(m.from_user and m.from_user.id in _WAITING_IMPORT_PAYS),
)
async def import_pays_document(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        _WAITING_IMPORT_PAYS.discard(message.from_user.id)
        return

    doc = message.document
    if not doc or not doc.file_name:
        await message.answer("Пришлите файл с расширением .xlsx")
        return
    if not doc.file_name.lower().endswith(".xlsx"):
        await message.answer("Нужен файл .xlsx (Excel).")
        return

    if doc.file_size is not None and doc.file_size > TELEGRAM_BOT_MAX_DOWNLOAD_BYTES:
        _WAITING_IMPORT_PAYS.discard(message.from_user.id)
        await message.answer(
            "❌ Файл больше <b>20 МБ</b>.",
            parse_mode="HTML",
        )
        return

    _WAITING_IMPORT_PAYS.discard(message.from_user.id)
    await message.answer("⏳ Обрабатываю платежи и обновляю подписки…")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    try:
        await message.bot.download(doc, destination=tmp_path)
        rows, skipped_parse = await asyncio.to_thread(_parse_import_pays_rows, tmp_path)
        now_naive = _naive_utc(datetime.now(timezone.utc))
        log_lines: List[str] = []
        casual_ok = 0
        casual_unknown = 0
        white_ok = 0
        failed = 0

        for uid, amount in rows:
            if amount == 1 or amount == 399:
                continue
            user = await sql.get_user_object_by_user_id(uid)
            if not user:
                log_lines.append(f"⚠️ casual: нет user_id={uid} amount={amount}")
                failed += 1
                continue
            days = _casual_days_by_amount(amount)
            if days is None:
                casual_unknown += 1
                log_lines.append(f"⚠️ casual: неизвестная сумма user_id={uid} amount={amount}")
                continue
            try:
                sub_end = user.subscription_end_date
                if sub_end:
                    new_end = sub_end + timedelta(days=days)
                    note = "продление"
                else:
                    new_end = now_naive + timedelta(days=days)
                    await sql.update_subscribtion(uid, x3.generate_client_id(uid))
                    note = "с нуля + subscribtion"
                await sql.update_subscription_end_date(uid, new_end)
                casual_ok += 1
                log_lines.append(
                    f"✅ casual user_id={uid} amount={amount} +{days}d → "
                    f"{new_end.strftime('%Y-%m-%d %H:%M')} ({note})"
                )
            except Exception as e:
                failed += 1
                logger.exception("import_pays casual user_id=%s", uid)
                log_lines.append(f"❌ casual user_id={uid} amount={amount}: {e}")

        for uid, amount in rows:
            if amount != 399:
                continue
            user = await sql.get_user_object_by_user_id(uid)
            if not user:
                log_lines.append(f"⚠️ white: нет user_id={uid} amount=399")
                failed += 1
                continue
            try:
                we = user.white_subscription_end_date
                if we:
                    new_w = we + timedelta(days=30)
                    note = "продление +30д"
                else:
                    new_w = now_naive + timedelta(days=30)
                    await sql.update_white_subscription(uid, x3.generate_client_id(uid * 100))
                    note = "с нуля + white_subscription"
                await sql.update_white_subscription_end_date(uid, new_w)
                white_ok += 1
                log_lines.append(
                    f"✅ white user_id={uid} amount=399 → "
                    f"{new_w.strftime('%Y-%m-%d %H:%M')} ({note})"
                )
            except Exception as e:
                failed += 1
                logger.exception("import_pays white user_id=%s", uid)
                log_lines.append(f"❌ white user_id={uid}: {e}")

        await _flush_import_pays_log(message, log_lines)
        await message.answer(
            "📊 <b>Итог import_pays</b>\n"
            f"Обычные тарифы применено: {casual_ok}\n"
            f"White 399 применено: {white_ok}\n"
            f"Строк без User ID/Amount: {skipped_parse}\n"
            f"Неизвестная сумма (casual): {casual_unknown}\n"
            f"Ошибок / нет пользователя: {failed}",
            parse_mode="HTML",
        )
        logger.info(
            "Админ %s import_pays: casual_ok=%s white_ok=%s skipped=%s unknown=%s failed=%s",
            message.from_user.id,
            casual_ok,
            white_ok,
            skipped_parse,
            casual_unknown,
            failed,
        )
    except TelegramBadRequest as e:
        err = str(e).lower()
        logger.exception("import_pays Telegram error")
        if "too big" in err or "file is too large" in err:
            await message.answer("❌ Файл слишком большой для скачивания ботом (~20 МБ).")
        else:
            await message.answer(f"❌ Telegram: {e}")
    except Exception as e:
        logger.exception("import_pays failed")
        await message.answer(f"❌ Ошибка: {e}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@router.message(
    F.document,
    lambda m: bool(m.from_user and m.from_user.id in _WAITING_IMPORT_EXCEL),
)
async def import_excel_document(message: Message) -> None:
    if not message.from_user or message.from_user.id not in ADMIN_IDS:
        _WAITING_IMPORT_EXCEL.discard(message.from_user.id)
        return

    doc = message.document
    if not doc or not doc.file_name:
        await message.answer("Пришлите файл с расширением .xlsx")
        return
    if not doc.file_name.lower().endswith(".xlsx"):
        await message.answer("Нужен файл .xlsx (Excel).")
        return

    if doc.file_size is not None and doc.file_size > TELEGRAM_BOT_MAX_DOWNLOAD_BYTES:
        _WAITING_IMPORT_EXCEL.discard(message.from_user.id)
        await message.answer(
            "❌ Файл больше <b>20 МБ</b> — Telegram не отдаёт такие файлы боту через API.\n\n"
            "Скопируйте .xlsx на VPS (scp / WinSCP) и выполните на сервере из каталога проекта:\n"
            "<code>python scripts/import_export_xlsx.py /полный/путь/к/файлу.xlsx</code>\n\n"
            "Либо поднимите <a href=\"https://core.telegram.org/bots/api#using-a-local-bot-api-server\">локальный Bot API server</a> "
            "и укажите его в клиенте бота — тогда лимит выше.",
            parse_mode="HTML",
            disable_web_page_preview=True,
        )
        return

    _WAITING_IMPORT_EXCEL.discard(message.from_user.id)
    await message.answer("⏳ Загружаю и разбираю файл…")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    try:
        await message.bot.download(doc, destination=tmp_path)
        bundles = await asyncio.to_thread(_parse_workbook, tmp_path)
        stats = await sql.import_replace_all_from_export_workbook(**bundles)
        body = "\n".join(f"{k}: {v}" for k, v in sorted(stats.items()))
        await message.answer("✅ Импорт завершён.\n" + body)
        logger.info("Админ %s импортировал Excel: %s", message.from_user.id, stats)
    except TelegramBadRequest as e:
        err = str(e).lower()
        logger.exception("import_excel Telegram error")
        if "too big" in err or "file is too large" in err:
            await message.answer(
                "❌ Telegram: файл слишком большой для скачивания ботом (лимит API ~20 МБ).\n\n"
                "На сервере: <code>python scripts/import_export_xlsx.py /путь/к/export.xlsx</code>",
                parse_mode="HTML",
            )
        else:
            await message.answer(f"❌ Telegram: {e}")
    except Exception as e:
        logger.exception("import_excel failed")
        await message.answer(f"❌ Ошибка импорта: {e}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
